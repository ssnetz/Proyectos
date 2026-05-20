"""
Genera el archivo SQL para importar la base de personas desde el Excel.
Uso: python3 import_personas.py [ruta_excel]
Salida: personas_import.sql (ejecutar en MySQL)
"""
import sys
import openpyxl
import re
import os

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), '../../../Base_Personas.xlsx')
OUTPUT_SQL  = os.path.join(os.path.dirname(__file__), 'personas_import.sql')
BATCH_SIZE  = 500

def esc(v):
    if v is None or str(v).strip() == '':
        return 'NULL'
    s = str(v).strip().replace('\\', '\\\\').replace("'", "\\'")
    return f"'{s}'"

print(f'Leyendo {EXCEL_PATH}...')
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws = wb.active

headers = [str(c.value).strip().upper() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
col = {h: i for i, h in enumerate(headers)}

rows_written = 0
batch = []

with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write("-- Importación de personas\n")
    f.write("-- Ejecutar en MySQL: mysql -u root -p stock_control < personas_import.sql\n\n")
    f.write("SET NAMES utf8mb4;\n")
    f.write("SET FOREIGN_KEY_CHECKS=0;\n\n")

    def flush(batch, f):
        if not batch:
            return
        f.write(
            "INSERT IGNORE INTO personas "
            "(tipo_documento, documento, apellido, nombre, sexo, calle, numeracion, departamento, piso, barrio, cuit_cuil) VALUES\n"
        )
        f.write(',\n'.join(batch) + ';\n\n')

    for row in ws.iter_rows(min_row=2, values_only=True):
        tipo   = str(row[col.get('TIPO_DOCUMENTO', 1)] or '1').strip()
        doc    = str(row[col.get('DOCUMENTO', 2)] or '').strip()
        apell  = str(row[col.get('APELLIDO', 3)] or '').strip()
        nombre = str(row[col.get('NOMBRE', 4)] or '').strip()
        sexo   = str(row[col.get('SEXO', 5)] or '').strip()[:1]
        calle  = str(row[col.get('CALLE_NOCOD', 6)] or '').strip()
        numer  = str(row[col.get('NUMERACION_CALLE', 7)] or '').strip()
        dpto   = str(row[col.get('DEPARTAMENTO', 8)] or '').strip()
        piso   = str(row[col.get('PISO', 9)] or '').strip()
        barrio = str(row[col.get('BARRIO', 10)] or '').strip()
        cuit   = str(row[col.get('CUIT_CUIL', 11)] or '').strip()

        if not doc or not apell:
            continue

        batch.append(
            f"({esc(tipo)},{esc(doc)},{esc(apell)},{esc(nombre)},{esc(sexo) if sexo else 'NULL'},"
            f"{esc(calle)},{esc(numer)},{esc(dpto)},{esc(piso)},{esc(barrio)},{esc(cuit)})"
        )
        rows_written += 1

        if len(batch) >= BATCH_SIZE:
            flush(batch, f)
            batch = []
            print(f'  {rows_written:,} filas procesadas...', end='\r')

    flush(batch, f)
    f.write("SET FOREIGN_KEY_CHECKS=1;\n")

wb.close()
print(f'\nTotal importadas: {rows_written:,} personas')
print(f'SQL generado en: {OUTPUT_SQL}')
print(f'\nPara importar ejecuta:')
print(f'  mysql -u root -p stock_control < {OUTPUT_SQL}')
