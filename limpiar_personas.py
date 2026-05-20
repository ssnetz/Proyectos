import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = '/root/.claude/uploads/797da9ef-038e-430a-9db8-dd275c2c14cb/66c3ba92-personas.csv'
OUTPUT = '/home/user/Proyectos/personas_limpias.xlsx'

HEADER_FILL   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT   = Font(color='FFFFFF', bold=True)
CUIT_FILL     = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
DUP_FILL      = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
THIN_BORDER   = Border(
    bottom=Side(style='thin', color='CCCCCC')
)

import re
_ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def sanitize(v):
    if isinstance(v, str):
        return _ILLEGAL.sub('', v)
    return v

def cuit_to_dni(doc):
    """Quita los 2 primeros dígitos y el último dígito de un CUIT de 11 dígitos."""
    return str(int(doc[2:-1]))   # int() elimina ceros a la izquierda

def write_header(ws, headers, col_widths=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def set_row(ws, row_idx, values, fill=None):
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=sanitize(v))
        cell.alignment = Alignment(vertical='center')
        if fill:
            cell.fill = fill

# ── 1. Leer CSV ──────────────────────────────────────────────────────────────
print('Leyendo CSV...')
rows = []
with open(INPUT, encoding='utf-8-sig') as f:
    reader = csv.DictReader(f, delimiter=';')
    fieldnames = reader.fieldnames
    for row in reader:
        rows.append({k: (v or '').strip() for k, v in row.items()})

print(f'  Total filas: {len(rows):,}')

# ── 2. Convertir CUITs a DNI ─────────────────────────────────────────────────
cuit_changes = []
for i, row in enumerate(rows):
    tipo = row['TIPO_DOCUMENTO']
    doc  = row['DOCUMENTO']
    if tipo == '5' and len(doc) == 11 and doc.isdigit():
        nuevo_doc = cuit_to_dni(doc)
        cuit_changes.append({
            'N_FILA'      : i + 2,
            'ID'          : row['ID'],
            'APELLIDO'    : row['APELLIDO'],
            'NOMBRE'      : row['NOMBRE'],
            'DOC_ORIGINAL': doc,
            'DOC_NUEVO'   : nuevo_doc,
            'TIPO_ORIG'   : '5 (CUIT)',
            'TIPO_NUEVO'  : '1 (DNI)',
        })
        row['DOCUMENTO']      = nuevo_doc
        row['TIPO_DOCUMENTO'] = '1'

print(f'  CUITs convertidos a DNI: {len(cuit_changes):,}')

# ── 3. Eliminar duplicados (por DOCUMENTO, manteniendo primera aparición) ────
seen   = {}    # doc -> índice en rows
clean  = []
dup_changes = []

for i, row in enumerate(rows):
    doc = row['DOCUMENTO']
    if doc not in seen:
        seen[doc] = i
        clean.append(row)
    else:
        dup_changes.append({
            'N_FILA'        : i + 2,
            'ID'            : row['ID'],
            'APELLIDO'      : row['APELLIDO'],
            'NOMBRE'        : row['NOMBRE'],
            'DOCUMENTO'     : doc,
            'TIPO_DOCUMENTO': row['TIPO_DOCUMENTO'],
            'FILA_MANTENIDA': seen[doc] + 2,
            'ID_MANTENIDO'  : rows[seen[doc]]['ID'],
        })

print(f'  Duplicados eliminados:  {len(dup_changes):,}')
print(f'  Filas finales limpias:  {len(clean):,}')

# ── 4. Crear Excel ───────────────────────────────────────────────────────────
print('Creando Excel...')
wb = Workbook()

# ── Pestaña 1: DATOS_LIMPIOS ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'DATOS_LIMPIOS'
ws1.freeze_panes = 'A2'

col_widths_data = [8, 8, 14, 30, 30, 6, 30, 10, 6, 6, 20, 14, 45]
write_header(ws1, fieldnames, col_widths_data)

for r, row in enumerate(clean, 2):
    vals = [row[f] for f in fieldnames]
    set_row(ws1, r, vals)
    if r % 2 == 0:
        for c in range(1, len(fieldnames) + 1):
            ws1.cell(row=r, column=c).fill = PatternFill(
                start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

print(f'  Hoja DATOS_LIMPIOS: {len(clean):,} filas')

# ── Pestaña 2: CAMBIOS ───────────────────────────────────────────────────────
ws2 = wb.create_sheet('CAMBIOS')
ws2.freeze_panes = 'A2'

# Sección A: conversiones CUIT → DNI
cuit_headers = [
    'TIPO_CAMBIO', 'N_FILA_CSV', 'ID', 'APELLIDO', 'NOMBRE',
    'DOCUMENTO_ORIGINAL (CUIT)', 'DOCUMENTO_NUEVO (DNI)',
    'TIPO_DOC_ORIGINAL', 'TIPO_DOC_NUEVO',
]
cuit_widths = [20, 12, 10, 30, 30, 26, 22, 18, 14]
write_header(ws2, cuit_headers, cuit_widths)

current_row = 2
for ch in cuit_changes:
    vals = [
        'Conversión CUIT → DNI',
        ch['N_FILA'], ch['ID'], ch['APELLIDO'], ch['NOMBRE'],
        ch['DOC_ORIGINAL'], ch['DOC_NUEVO'],
        ch['TIPO_ORIG'], ch['TIPO_NUEVO'],
    ]
    set_row(ws2, current_row, vals, fill=CUIT_FILL)
    current_row += 1

# Separador
current_row += 1

# Encabezado sección B: duplicados
dup_headers = [
    'TIPO_CAMBIO', 'N_FILA_CSV', 'ID', 'APELLIDO', 'NOMBRE',
    'DOCUMENTO', 'TIPO_DOC',
    'FILA_MANTENIDA', 'ID_MANTENIDO',
]
for c, h in enumerate(dup_headers, 1):
    cell = ws2.cell(row=current_row, column=c, value=h)
    cell.fill = PatternFill(start_color='833C00', end_color='833C00', fill_type='solid')
    cell.font = Font(color='FFFFFF', bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[current_row].height = 22
current_row += 1

for dc in dup_changes:
    vals = [
        'Duplicado eliminado',
        dc['N_FILA'], dc['ID'], dc['APELLIDO'], dc['NOMBRE'],
        dc['DOCUMENTO'], dc['TIPO_DOCUMENTO'],
        dc['FILA_MANTENIDA'], dc['ID_MANTENIDO'],
    ]
    set_row(ws2, current_row, vals, fill=DUP_FILL)
    current_row += 1

print(f'  Hoja CAMBIOS: {len(cuit_changes):,} conversiones + {len(dup_changes):,} duplicados')

# ── Pestaña 3: RESUMEN ───────────────────────────────────────────────────────
ws3 = wb.create_sheet('RESUMEN')
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 18

summary_data = [
    ('RESUMEN DE CAMBIOS', ''),
    ('', ''),
    ('Total filas originales',            len(rows)),
    ('CUITs convertidos a DNI',           len(cuit_changes)),
    ('Duplicados eliminados',             len(dup_changes)),
    ('Total filas en DATOS_LIMPIOS',      len(clean)),
    ('', ''),
    ('Filas eliminadas / originales (%)', f'{len(dup_changes)/len(rows)*100:.2f}%'),
]

for r, (label, value) in enumerate(summary_data, 1):
    cell_a = ws3.cell(row=r, column=1, value=label)
    cell_b = ws3.cell(row=r, column=2, value=value)
    if r == 1:
        cell_a.font = Font(bold=True, size=14, color='1F4E79')
    elif label and label != '':
        cell_a.font = Font(bold=False)
        if isinstance(value, int):
            cell_b.alignment = Alignment(horizontal='right')
            cell_b.font = Font(bold=True, color='1F4E79')

# ── Guardar ──────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f'\nArchivo guardado: {OUTPUT}')
print('Listo.')
